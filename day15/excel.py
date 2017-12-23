import openpyxl
from pprint import pprint

# with openpyxl.load_workbook("example.xlsx") as plik:
#     arkusze = plik.sheetnames
#     print(arkusze)
# exit()

plik = openpyxl.load_workbook("example.xlsx")

arkusze = plik.sheetnames
print(arkusze)

aktywny_arkusz = plik.get_sheet_by_name("Owoce")
print(f"Aktywny arkusz: {aktywny_arkusz.title}")

# komórki
komorka = aktywny_arkusz["A1"]
print(komorka)
print(komorka.value)

# koordynaty
print(f"Adres komórki: {komorka.coordinate}")
print(f"Kolumna komórki: {komorka.column}")
print(f"Wiersz komórki: {komorka.row}")

print(aktywny_arkusz.cell(row=2, column=2))

# zamiana liter kolumn na liczbę i vice versa
from openpyxl.utils import get_column_letter, column_index_from_string

print(get_column_letter(123))
print(column_index_from_string("ABC"))

# maksymalny rozmiar
print("Ostatnia kolumna:")
print(aktywny_arkusz.max_column)
print(get_column_letter(aktywny_arkusz.max_column))
print("Ostatni wiersz")
print(aktywny_arkusz.max_row)

pprint(aktywny_arkusz["A1":"C7"], indent=4)

for wiersz in aktywny_arkusz["A1":"C7"]:
    for kom in wiersz:
        print(kom.value, end='   ')
    print()

aktywny_arkusz["B9"] = "hello from Python"
plik.save("example2.xlsx")
plik.close()
